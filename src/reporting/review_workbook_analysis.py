from __future__ import annotations

import shutil
import tempfile
from pathlib import Path
from typing import Iterable, Optional

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string


SOURCE_SHEET_NAME = "Sheet1"
ANALYSIS_SHEET_NAME = "Analysis_Data"
KPI_SHEET_NAME = "KPI_Summary"
CHART_SHEET_NAME = "Charts"

YEAR_TOPIC_COUNT_SHEET = "Year_Topic_Count"
YEAR_TOPIC_SHARE_SHEET = "Year_Topic_Share"
YEAR_LEVEL_COUNT_SHEET = "Year_Level_Count"
YEAR_SPATIAL_FLAG_SHEET = "Year_SpatialFlag_Count"
LEVEL_SHARE_SHEET = "Level_Share"
LEVEL_TOPIC_COUNT_SHEET = "Level_Topic_Count"
LEVEL_TOPIC_SHARE_SHEET = "Level_Topic_Share"
SPACE_COUNT_SHEET = "Space_Count"
SPACE_TOPIC_COUNT_SHEET = "Space_Topic_Count"
SPACE_TOPIC_SHARE_SHEET = "Space_Topic_Share"
TOPIC_SPACE_TOP_SHEET = "Topic_Space_Top"

ANALYSIS_SHEETS = [
    ANALYSIS_SHEET_NAME,
    KPI_SHEET_NAME,
    YEAR_TOPIC_COUNT_SHEET,
    YEAR_TOPIC_SHARE_SHEET,
    YEAR_LEVEL_COUNT_SHEET,
    YEAR_SPATIAL_FLAG_SHEET,
    LEVEL_SHARE_SHEET,
    LEVEL_TOPIC_COUNT_SHEET,
    LEVEL_TOPIC_SHARE_SHEET,
    SPACE_COUNT_SHEET,
    SPACE_TOPIC_COUNT_SHEET,
    SPACE_TOPIC_SHARE_SHEET,
    TOPIC_SPACE_TOP_SHEET,
    CHART_SHEET_NAME,
]

COL_YEAR = "Publication Year"
COL_TOPIC = "topic_final"
COL_TOPIC_NAME_EN = "topic_final_name_en"
COL_TOPIC_NAME_ZH = "topic_final_name_zh"
COL_SPATIAL_FLAG = "\u9884\u6d4b_\u7a7a\u95f4\u7814\u7a76/\u975e\u7a7a\u95f4\u7814\u7a76"
COL_SPATIAL_LEVEL = "\u9884\u6d4b_\u7a7a\u95f4\u7b49\u7ea7"
COL_SPACE_DETAIL = "\u9884\u6d4b_\u5177\u4f53\u7a7a\u95f4\u63cf\u8ff0"

ANALYSIS_YEAR = "analysis_publication_year"
ANALYSIS_YEAR_VALID = "analysis_has_valid_year"
ANALYSIS_SPATIAL_FLAG = "analysis_spatial_flag"
ANALYSIS_IS_SPATIAL = "analysis_is_spatial"
ANALYSIS_SPATIAL_LEVEL = "analysis_spatial_level"
ANALYSIS_SPACE_DETAIL = "analysis_space_detail"
ANALYSIS_HAS_SPACE_DETAIL = "analysis_has_space_detail"

NOT_MENTIONED = "Not mentioned"
SPATIAL = "Spatial"
NON_SPATIAL = "Non-spatial"
TOTAL_LABEL = "Total"

LEVEL_NORMALIZATION = {
    "9. Micro / Neighborhood / Block": "9. Micro / Neighborhood / Block Scale",
}

DISPLAY_COLUMN_NAMES = {
    ANALYSIS_YEAR: COL_YEAR,
    ANALYSIS_SPATIAL_LEVEL: COL_SPATIAL_LEVEL,
    ANALYSIS_SPACE_DETAIL: COL_SPACE_DETAIL,
}

COLORS = {
    "ink": "#17324D",
    "grid": "#D9E3F0",
    "blue": "#4C78A8",
    "teal": "#54A24B",
    "orange": "#F58518",
    "red": "#C44E52",
    "purple": "#B279A2",
    "gold": "#ECA82C",
    "slate": "#6E7C8C",
}

HEADER_FILL = PatternFill("solid", fgColor="DCEAF7")
SECTION_FILL = PatternFill("solid", fgColor="EDF4FB")
HEADER_FONT = Font(bold=True)
WRAP = Alignment(vertical="top", wrap_text=True)

CHART_LAYOUTS = [
    ("Year-Topic Share", "A3", 760, 360),
    ("Year-Topic Count Heatmap", "K3", 760, 430),
    ("Spatial Level Share", "A26", 760, 360),
    ("Spatial Level vs Topic", "K26", 760, 400),
    ("Top Detailed Spaces", "A49", 760, 400),
    ("Year vs Spatial Level", "K49", 760, 430),
]


def configure_matplotlib() -> None:
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["figure.facecolor"] = "white"
    plt.rcParams["axes.facecolor"] = "white"


def load_review_sheet(workbook_path: Path, sheet_name: str = SOURCE_SHEET_NAME) -> pd.DataFrame:
    return pd.read_excel(workbook_path, sheet_name=sheet_name, engine="openpyxl")


def normalize_review_frame(df: pd.DataFrame) -> pd.DataFrame:
    working = df.copy()

    if COL_YEAR not in working.columns:
        working[COL_YEAR] = pd.NA
    years = pd.to_numeric(working[COL_YEAR], errors="coerce")
    working[ANALYSIS_YEAR] = years.astype("Int64")
    working[ANALYSIS_YEAR_VALID] = working[ANALYSIS_YEAR].notna()

    if COL_TOPIC_NAME_ZH not in working.columns:
        working[COL_TOPIC_NAME_ZH] = ""
    if COL_TOPIC_NAME_EN not in working.columns:
        working[COL_TOPIC_NAME_EN] = ""
    if COL_TOPIC not in working.columns:
        working[COL_TOPIC] = ""

    working[COL_TOPIC_NAME_ZH] = working[COL_TOPIC_NAME_ZH].fillna("").astype(str).str.strip()
    working[COL_TOPIC_NAME_EN] = working[COL_TOPIC_NAME_EN].fillna("").astype(str).str.strip()
    working[COL_TOPIC] = working[COL_TOPIC].fillna("").astype(str).str.strip()

    working[ANALYSIS_SPATIAL_FLAG] = working.get(COL_SPATIAL_FLAG, "").apply(_normalize_spatial_flag)
    working[ANALYSIS_IS_SPATIAL] = working[ANALYSIS_SPATIAL_FLAG].eq(SPATIAL)

    level_series = working.get(COL_SPATIAL_LEVEL, "")
    working[ANALYSIS_SPATIAL_LEVEL] = (
        pd.Series(level_series, index=working.index)
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", NOT_MENTIONED)
        .replace(LEVEL_NORMALIZATION)
    )

    detail_series = working.get(COL_SPACE_DETAIL, "")
    working[ANALYSIS_SPACE_DETAIL] = (
        pd.Series(detail_series, index=working.index)
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", NOT_MENTIONED)
    )
    working[ANALYSIS_HAS_SPACE_DETAIL] = working[ANALYSIS_SPACE_DETAIL].ne(NOT_MENTIONED)

    return working


def build_analysis_tables(analysis_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    year_valid_df = analysis_df[analysis_df[ANALYSIS_YEAR_VALID]].copy()
    spatial_df = analysis_df[analysis_df[ANALYSIS_IS_SPATIAL]].copy()
    detailed_space_df = spatial_df[spatial_df[ANALYSIS_HAS_SPACE_DETAIL]].copy()

    topic_order = _ordered_values(analysis_df[COL_TOPIC_NAME_ZH])
    level_order = _ordered_values(spatial_df[ANALYSIS_SPATIAL_LEVEL])

    tables: dict[str, pd.DataFrame] = {}

    analysis_cols = list(analysis_df.columns)
    preferred_tail = [
        ANALYSIS_YEAR,
        ANALYSIS_YEAR_VALID,
        ANALYSIS_SPATIAL_FLAG,
        ANALYSIS_IS_SPATIAL,
        ANALYSIS_SPATIAL_LEVEL,
        ANALYSIS_SPACE_DETAIL,
        ANALYSIS_HAS_SPACE_DETAIL,
    ]
    analysis_cols = [column for column in analysis_cols if column not in preferred_tail] + preferred_tail
    tables[ANALYSIS_SHEET_NAME] = analysis_df[analysis_cols].copy()

    year_topic_count = _count_pivot(
        year_valid_df,
        index=ANALYSIS_YEAR,
        columns=COL_TOPIC_NAME_ZH,
        column_order=topic_order,
        sort_index=True,
    )
    tables[YEAR_TOPIC_COUNT_SHEET] = _with_totals(year_topic_count)
    tables[YEAR_TOPIC_SHARE_SHEET] = _share_table(year_topic_count)

    year_level_count = _count_pivot(
        year_valid_df,
        index=ANALYSIS_YEAR,
        columns=ANALYSIS_SPATIAL_LEVEL,
        column_order=_ordered_values(year_valid_df[ANALYSIS_SPATIAL_LEVEL]),
        sort_index=True,
    )
    tables[YEAR_LEVEL_COUNT_SHEET] = _with_totals(year_level_count)

    year_spatial = (
        year_valid_df.groupby([ANALYSIS_YEAR, ANALYSIS_SPATIAL_FLAG], dropna=False)
        .size()
        .unstack(fill_value=0)
        .reindex(columns=[SPATIAL, NON_SPATIAL], fill_value=0)
        .sort_index()
    )
    year_spatial_summary = year_spatial.reset_index().rename(columns={ANALYSIS_YEAR: COL_YEAR})
    year_spatial_summary["Total_Count"] = year_spatial_summary[[SPATIAL, NON_SPATIAL]].sum(axis=1)
    year_spatial_summary["Spatial_Share"] = np.where(
        year_spatial_summary["Total_Count"].gt(0),
        year_spatial_summary[SPATIAL] / year_spatial_summary["Total_Count"],
        0.0,
    )
    year_spatial_summary["Non-spatial_Share"] = np.where(
        year_spatial_summary["Total_Count"].gt(0),
        year_spatial_summary[NON_SPATIAL] / year_spatial_summary["Total_Count"],
        0.0,
    )
    total_row = {
        COL_YEAR: TOTAL_LABEL,
        SPATIAL: int(year_spatial_summary[SPATIAL].sum()),
        NON_SPATIAL: int(year_spatial_summary[NON_SPATIAL].sum()),
    }
    total_row["Total_Count"] = total_row[SPATIAL] + total_row[NON_SPATIAL]
    total_row["Spatial_Share"] = total_row[SPATIAL] / total_row["Total_Count"] if total_row["Total_Count"] else 0.0
    total_row["Non-spatial_Share"] = (
        total_row[NON_SPATIAL] / total_row["Total_Count"] if total_row["Total_Count"] else 0.0
    )
    tables[YEAR_SPATIAL_FLAG_SHEET] = pd.concat(
        [year_spatial_summary, pd.DataFrame([total_row])],
        ignore_index=True,
    )

    level_counts = spatial_df[ANALYSIS_SPATIAL_LEVEL].value_counts().reindex(level_order, fill_value=0)
    spatial_total = int(spatial_df.shape[0])
    all_total = int(analysis_df.shape[0])
    level_share = pd.DataFrame(
        {
            COL_SPATIAL_LEVEL: level_counts.index,
            "Count": level_counts.values,
            "Share_All_Samples": np.where(all_total, level_counts.values / all_total, 0.0),
            "Share_Spatial_Samples": np.where(spatial_total, level_counts.values / spatial_total, 0.0),
        }
    )
    tables[LEVEL_SHARE_SHEET] = level_share

    level_topic_count = _count_pivot(
        spatial_df,
        index=ANALYSIS_SPATIAL_LEVEL,
        columns=COL_TOPIC_NAME_ZH,
        column_order=topic_order,
        index_order=level_order,
    )
    tables[LEVEL_TOPIC_COUNT_SHEET] = _with_totals(level_topic_count)
    tables[LEVEL_TOPIC_SHARE_SHEET] = _share_table(level_topic_count)

    space_group = detailed_space_df.groupby(ANALYSIS_SPACE_DETAIL, dropna=False)
    space_count = (
        space_group.size()
        .rename("Count")
        .sort_values(ascending=False)
        .reset_index()
        .rename(columns={ANALYSIS_SPACE_DETAIL: COL_SPACE_DETAIL})
    )
    detailed_total = int(detailed_space_df.shape[0])
    space_count["Overall_Rank"] = np.arange(1, len(space_count) + 1)
    space_count["Share_All_Samples"] = np.where(all_total, space_count["Count"] / all_total, 0.0)
    space_count["Share_Detailed_Spatial"] = np.where(detailed_total, space_count["Count"] / detailed_total, 0.0)
    if detailed_space_df[ANALYSIS_YEAR_VALID].any():
        year_summary = (
            detailed_space_df[detailed_space_df[ANALYSIS_YEAR_VALID]]
            .groupby(ANALYSIS_SPACE_DETAIL)[ANALYSIS_YEAR]
            .agg(["min", "max"])
            .rename(columns={"min": "First_Year", "max": "Last_Year"})
            .reset_index()
            .rename(columns={ANALYSIS_SPACE_DETAIL: COL_SPACE_DETAIL})
        )
        space_count = space_count.merge(year_summary, on=COL_SPACE_DETAIL, how="left")
    else:
        space_count["First_Year"] = pd.NA
        space_count["Last_Year"] = pd.NA
    tables[SPACE_COUNT_SHEET] = space_count

    space_topic_count = _count_pivot(
        detailed_space_df,
        index=ANALYSIS_SPACE_DETAIL,
        columns=COL_TOPIC_NAME_ZH,
        column_order=topic_order,
        index_order=space_count[COL_SPACE_DETAIL].tolist(),
    )
    tables[SPACE_TOPIC_COUNT_SHEET] = _with_totals(space_topic_count)
    tables[SPACE_TOPIC_SHARE_SHEET] = _share_table(space_topic_count)

    space_rank_lookup = dict(zip(space_count[COL_SPACE_DETAIL], space_count["Overall_Rank"]))
    topic_space_rows = []
    for topic_name, topic_df in detailed_space_df.groupby(COL_TOPIC_NAME_ZH, dropna=False):
        topic_total = int(topic_df.shape[0])
        grouped = (
            topic_df.groupby(ANALYSIS_SPACE_DETAIL, dropna=False)
            .size()
            .rename("Count")
            .sort_values(ascending=False)
            .reset_index()
        )
        for _, row in grouped.head(10).iterrows():
            topic_space_rows.append(
                {
                    COL_TOPIC_NAME_ZH: topic_name,
                    COL_SPACE_DETAIL: row[ANALYSIS_SPACE_DETAIL],
                    "Count": int(row["Count"]),
                    "Share_Within_Topic": row["Count"] / topic_total if topic_total else 0.0,
                    "Topic_Total_Count": topic_total,
                    "Overall_Space_Rank": int(space_rank_lookup.get(row[ANALYSIS_SPACE_DETAIL], 0)),
                }
            )
    tables[TOPIC_SPACE_TOP_SHEET] = pd.DataFrame(topic_space_rows)

    return tables


def generate_review_analysis_workbook(
    input_path: Path,
    *,
    append: bool = True,
    replace_analysis_sheets: bool = False,
    output_path: Optional[Path] = None,
) -> Path:
    input_path = Path(input_path)
    if append:
        target_path = input_path
    else:
        if output_path is None:
            raise ValueError("output_path is required when append is False")
        target_path = Path(output_path)
        if input_path.resolve() != target_path.resolve():
            shutil.copy2(input_path, target_path)

    source_df = load_review_sheet(target_path)
    analysis_df = normalize_review_frame(source_df)
    tables = build_analysis_tables(analysis_df)
    kpi_blocks = build_kpi_blocks(analysis_df, tables)

    _prepare_workbook(target_path, replace_analysis_sheets=replace_analysis_sheets)
    _write_analysis_tables(target_path, tables)
    _finalize_workbook(target_path, tables, kpi_blocks)
    return target_path


def build_kpi_blocks(analysis_df: pd.DataFrame, tables: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    spatial_total = int(analysis_df[ANALYSIS_IS_SPATIAL].sum())
    non_spatial_total = int((~analysis_df[ANALYSIS_IS_SPATIAL]).sum())
    valid_year_df = analysis_df[analysis_df[ANALYSIS_YEAR_VALID]]
    detailed_space_df = analysis_df[analysis_df[ANALYSIS_HAS_SPACE_DETAIL] & analysis_df[ANALYSIS_IS_SPATIAL]]

    summary = pd.DataFrame(
        [
            {"Metric": "Total Samples", "Value": int(len(analysis_df))},
            {
                "Metric": "Year Range",
                "Value": (
                    f"{int(valid_year_df[ANALYSIS_YEAR].min())}-{int(valid_year_df[ANALYSIS_YEAR].max())}"
                    if not valid_year_df.empty
                    else ""
                ),
            },
            {"Metric": "Distinct Topics", "Value": int(analysis_df[COL_TOPIC_NAME_ZH].nunique(dropna=True))},
            {"Metric": "Spatial Studies", "Value": spatial_total},
            {"Metric": "Non-spatial Studies", "Value": non_spatial_total},
            {
                "Metric": "Distinct Detailed Spaces",
                "Value": int(detailed_space_df[ANALYSIS_SPACE_DETAIL].nunique(dropna=True)),
            },
            {
                "Metric": "Not Mentioned Count",
                "Value": int(analysis_df[ANALYSIS_SPACE_DETAIL].eq(NOT_MENTIONED).sum()),
            },
            {
                "Metric": "Not Mentioned Share",
                "Value": (
                    analysis_df[ANALYSIS_SPACE_DETAIL].eq(NOT_MENTIONED).mean()
                    if len(analysis_df)
                    else 0.0
                ),
            },
        ]
    )

    top_topics = (
        analysis_df[COL_TOPIC_NAME_ZH]
        .value_counts()
        .head(10)
        .rename_axis(COL_TOPIC_NAME_ZH)
        .reset_index(name="Count")
    )
    top_topics["Share"] = np.where(len(analysis_df), top_topics["Count"] / len(analysis_df), 0.0)

    level_share = tables[LEVEL_SHARE_SHEET].head(10).copy()
    top_spaces = tables[SPACE_COUNT_SHEET].head(20).copy()

    return {
        "summary": summary,
        "top_topics": top_topics,
        "top_levels": level_share,
        "top_spaces": top_spaces,
    }


def _prepare_workbook(workbook_path: Path, *, replace_analysis_sheets: bool) -> None:
    workbook = load_workbook(workbook_path)
    existing = [sheet for sheet in ANALYSIS_SHEETS if sheet in workbook.sheetnames]
    if existing and not replace_analysis_sheets:
        raise ValueError(f"Analysis sheets already exist: {', '.join(existing)}")
    for sheet_name in existing:
        del workbook[sheet_name]
    workbook.save(workbook_path)


def _write_analysis_tables(workbook_path: Path, tables: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a") as writer:
        for sheet_name, table in tables.items():
            table.to_excel(writer, sheet_name=sheet_name, index=False)


def _finalize_workbook(workbook_path: Path, tables: dict[str, pd.DataFrame], kpi_blocks: dict[str, pd.DataFrame]) -> None:
    workbook = load_workbook(workbook_path)

    for sheet_name, table in tables.items():
        worksheet = workbook[sheet_name]
        percent_cols = _percent_columns_for_sheet(sheet_name, table)
        _style_table_sheet(worksheet, percent_columns=percent_cols)

    _write_kpi_sheet(workbook, kpi_blocks)

    with tempfile.TemporaryDirectory(prefix="review_analysis_figs_") as tmp_dir_name:
        tmp_dir = Path(tmp_dir_name)
        chart_paths = _create_chart_images(tables, tmp_dir)
        _write_chart_sheet(workbook, chart_paths)
        workbook.save(workbook_path)


def _percent_columns_for_sheet(sheet_name: str, table: pd.DataFrame) -> set[str]:
    explicit = {
        YEAR_SPATIAL_FLAG_SHEET: {"Spatial_Share", "Non-spatial_Share"},
        LEVEL_SHARE_SHEET: {"Share_All_Samples", "Share_Spatial_Samples"},
        SPACE_COUNT_SHEET: {"Share_All_Samples", "Share_Detailed_Spatial"},
        TOPIC_SPACE_TOP_SHEET: {"Share_Within_Topic"},
    }
    if sheet_name in explicit:
        return explicit[sheet_name]
    if sheet_name.endswith("_Share"):
        return set(table.columns[1:])
    return {column for column in table.columns if "Share" in str(column)}


def _style_table_sheet(worksheet, *, percent_columns: Iterable[str] = ()) -> None:
    percent_columns = set(percent_columns)
    if worksheet.max_row == 0 or worksheet.max_column == 0:
        return

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_lookup = {cell.column: str(cell.value) for cell in worksheet[1]}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            if header_lookup.get(cell.column) in percent_columns and isinstance(cell.value, (float, int)):
                cell.number_format = "0.0%"

    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells[: min(len(column_cells), 1000)]]
        max_len = max((len(value) for value in values), default=10)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 40)


def _write_kpi_sheet(workbook, kpi_blocks: dict[str, pd.DataFrame]) -> None:
    if KPI_SHEET_NAME in workbook.sheetnames:
        del workbook[KPI_SHEET_NAME]
    worksheet = workbook.create_sheet(KPI_SHEET_NAME)
    worksheet.freeze_panes = "A2"

    block_specs = [
        ("Summary Metrics", kpi_blocks["summary"], "A1"),
        ("Top Topics", kpi_blocks["top_topics"], "A14"),
        ("Top Levels", kpi_blocks["top_levels"], "F14"),
        ("Top Detailed Spaces", kpi_blocks["top_spaces"], "K14"),
    ]

    for title, frame, anchor in block_specs:
        start_col = worksheet[anchor].column
        start_row = worksheet[anchor].row
        title_cell = worksheet.cell(row=start_row, column=start_col, value=title)
        title_cell.fill = SECTION_FILL
        title_cell.font = HEADER_FONT
        title_cell.alignment = WRAP
        _write_dataframe_block(worksheet, frame, start_row=start_row + 1, start_col=start_col)

    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells[: min(len(column_cells), 200)]]
        max_len = max((len(value) for value in values), default=10)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 36)


def _write_dataframe_block(worksheet, frame: pd.DataFrame, *, start_row: int, start_col: int) -> None:
    for offset, column_name in enumerate(frame.columns):
        cell = worksheet.cell(row=start_row, column=start_col + offset, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP

    for row_offset, (_, row) in enumerate(frame.iterrows(), start=1):
        for col_offset, column_name in enumerate(frame.columns):
            value = row[column_name]
            if pd.isna(value):
                value = None
            cell = worksheet.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
            cell.alignment = WRAP
            if "Share" in str(column_name) and isinstance(row[column_name], (float, int)):
                cell.number_format = "0.0%"


def _write_chart_sheet(workbook, chart_paths: list[tuple[str, Path]]) -> None:
    if CHART_SHEET_NAME in workbook.sheetnames:
        del workbook[CHART_SHEET_NAME]
    worksheet = workbook.create_sheet(CHART_SHEET_NAME)
    worksheet.sheet_view.showGridLines = False
    worksheet["A1"] = "Review Workbook Visualization"
    worksheet["A1"].font = Font(size=14, bold=True)

    worksheet["A2"] = "Charts below summarize theme evolution by year, spatial level structure, and detailed space hotspots."
    worksheet["A2"].alignment = WRAP

    for (expected_title, image_path), (title, anchor, width, height) in zip(chart_paths, CHART_LAYOUTS):
        if expected_title != title:
            raise ValueError(f"Chart layout mismatch: {expected_title} != {title}")
        title_cell = worksheet[anchor]
        title_cell.value = title
        title_cell.font = HEADER_FONT
        title_cell.fill = SECTION_FILL
        image = XLImage(str(image_path))
        image.width = width
        image.height = height
        image.anchor = _image_anchor_from_title(anchor)
        worksheet.add_image(image)

    for column in range(1, 20):
        worksheet.column_dimensions[get_column_letter(column)].width = 12
    for row in range(1, 72):
        worksheet.row_dimensions[row].height = 22


def _image_anchor_from_title(anchor: str) -> str:
    column_letters, row = coordinate_from_string(anchor)
    next_row = row + 1
    return f"{column_letters}{next_row}"


def _create_chart_images(tables: dict[str, pd.DataFrame], figdir: Path) -> list[tuple[str, Path]]:
    configure_matplotlib()

    chart_paths = [
        ("Year-Topic Share", figdir / "year_topic_share.png"),
        ("Year-Topic Count Heatmap", figdir / "year_topic_count_heatmap.png"),
        ("Spatial Level Share", figdir / "level_share.png"),
        ("Spatial Level vs Topic", figdir / "level_topic_heatmap.png"),
        ("Top Detailed Spaces", figdir / "space_count_top20.png"),
        ("Year vs Spatial Level", figdir / "year_level_heatmap.png"),
    ]

    _save_year_topic_share_chart(tables[YEAR_TOPIC_SHARE_SHEET], chart_paths[0][1])
    _save_heatmap_chart(tables[YEAR_TOPIC_COUNT_SHEET], chart_paths[1][1], "Year vs Topic Count")
    _save_level_share_chart(tables[LEVEL_SHARE_SHEET], chart_paths[2][1])
    _save_heatmap_chart(tables[LEVEL_TOPIC_COUNT_SHEET], chart_paths[3][1], "Spatial Level vs Topic")
    _save_space_count_chart(tables[SPACE_COUNT_SHEET], chart_paths[4][1])
    _save_heatmap_chart(tables[YEAR_LEVEL_COUNT_SHEET], chart_paths[5][1], "Year vs Spatial Level")

    return chart_paths


def create_review_analysis_chart_images(tables: dict[str, pd.DataFrame], figdir: Path) -> list[tuple[str, Path]]:
    return _create_chart_images(tables, figdir)


def _save_year_topic_share_chart(frame: pd.DataFrame, output_path: Path) -> None:
    plot_df = frame[frame.iloc[:, 0].astype(str) != TOTAL_LABEL].copy()
    if plot_df.empty:
        return _save_placeholder_chart(output_path, "No year-topic share data")

    plot_df = plot_df.rename(columns={plot_df.columns[0]: COL_YEAR}).set_index(COL_YEAR)
    if TOTAL_LABEL in plot_df.columns:
        plot_df = plot_df.drop(columns=[TOTAL_LABEL])
    topic_totals = plot_df.sum(axis=0).sort_values(ascending=False)
    top_topics = topic_totals.head(10).index.tolist()
    chart_df = plot_df[top_topics].copy()
    other_columns = [column for column in plot_df.columns if column not in top_topics]
    if other_columns:
        chart_df["Other"] = plot_df[other_columns].sum(axis=1)

    years = chart_df.index.astype(int).tolist()
    fig, ax = plt.subplots(figsize=(11, 5.8))
    ax.stackplot(years, chart_df.T.values, labels=chart_df.columns, alpha=0.92)
    ax.set_title("Topic Share by Publication Year")
    ax.set_ylabel("Share")
    ax.set_xlabel("Publication Year")
    ax.set_ylim(0, 1)
    ax.yaxis.set_major_formatter(lambda value, _: f"{value:.0%}")
    ax.grid(axis="y", color=COLORS["grid"], linewidth=0.8, alpha=0.8)
    ax.legend(loc="upper left", bbox_to_anchor=(1.01, 1.02), frameon=False)
    fig.tight_layout()
    fig.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def _save_level_share_chart(frame: pd.DataFrame, output_path: Path) -> None:
    if frame.empty:
        return _save_placeholder_chart(output_path, "No spatial level share data")

    plot_df = frame.sort_values("Count", ascending=True)
    fig, axes = plt.subplots(1, 2, figsize=(12, 5.8), gridspec_kw={"width_ratios": [1.15, 1]})

    axes[0].barh(plot_df[COL_SPATIAL_LEVEL], plot_df["Count"], color=COLORS["blue"])
    axes[0].set_title("Spatial Level Count")
    axes[0].set_xlabel("Count")
    axes[0].grid(axis="x", color=COLORS["grid"], linewidth=0.8, alpha=0.8)

    share_positions = np.arange(len(plot_df))
    bar_height = 0.38
    axes[1].barh(
        share_positions - bar_height / 2,
        plot_df["Share_All_Samples"] * 100,
        height=bar_height,
        color=COLORS["teal"],
        label="All Samples",
    )
    axes[1].barh(
        share_positions + bar_height / 2,
        plot_df["Share_Spatial_Samples"] * 100,
        height=bar_height,
        color=COLORS["orange"],
        label="Spatial Samples",
    )
    axes[1].set_yticks(share_positions, plot_df[COL_SPATIAL_LEVEL])
    axes[1].set_title("Spatial Level Share")
    axes[1].set_xlabel("Share (%)")
    axes[1].legend(frameon=False, loc="lower right")
    axes[1].grid(axis="x", color=COLORS["grid"], linewidth=0.8, alpha=0.8)

    fig.tight_layout()
    fig.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def _save_space_count_chart(frame: pd.DataFrame, output_path: Path) -> None:
    plot_df = frame.head(20).copy()
    if plot_df.empty:
        return _save_placeholder_chart(output_path, "No detailed space data")

    plot_df = plot_df.sort_values("Count", ascending=True)
    fig, ax = plt.subplots(figsize=(12, 6.5))
    ax.barh(plot_df[COL_SPACE_DETAIL], plot_df["Count"], color=COLORS["purple"])
    ax.set_title("Top 20 Detailed Spaces")
    ax.set_xlabel("Count")
    ax.grid(axis="x", color=COLORS["grid"], linewidth=0.8, alpha=0.8)
    fig.tight_layout()
    fig.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def _save_heatmap_chart(frame: pd.DataFrame, output_path: Path, title: str) -> None:
    plot_df = frame.copy()
    if plot_df.empty:
        return _save_placeholder_chart(output_path, f"No data for {title}")

    first_col = plot_df.columns[0]
    plot_df = plot_df[plot_df[first_col].astype(str) != TOTAL_LABEL]
    if plot_df.empty:
        return _save_placeholder_chart(output_path, f"No data for {title}")

    plot_df = plot_df.rename(columns={first_col: "index_label"}).set_index("index_label")
    if TOTAL_LABEL in plot_df.columns:
        plot_df = plot_df.drop(columns=[TOTAL_LABEL])
    if TOTAL_LABEL in plot_df.index:
        plot_df = plot_df.drop(index=TOTAL_LABEL)

    if plot_df.empty or plot_df.shape[1] == 0:
        return _save_placeholder_chart(output_path, f"No data for {title}")

    if plot_df.shape[1] > 15:
        top_columns = plot_df.sum(axis=0).sort_values(ascending=False).head(15).index
        plot_df = plot_df.loc[:, top_columns]

    matrix = plot_df.to_numpy(dtype=float)
    fig_width = max(8.5, 0.45 * plot_df.shape[1] + 3.5)
    fig_height = max(4.5, 0.5 * plot_df.shape[0] + 2.5)
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    image = ax.imshow(matrix, cmap="YlGnBu", aspect="auto")
    ax.set_xticks(np.arange(plot_df.shape[1]), labels=plot_df.columns, rotation=35, ha="right")
    ax.set_yticks(np.arange(plot_df.shape[0]), labels=plot_df.index)
    ax.set_title(title)
    plt.colorbar(image, ax=ax, fraction=0.03, pad=0.02)
    fig.tight_layout()
    fig.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def _save_placeholder_chart(output_path: Path, message: str) -> None:
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.axis("off")
    ax.text(0.5, 0.5, message, ha="center", va="center", fontsize=14)
    fig.tight_layout()
    fig.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def _normalize_spatial_flag(value: object) -> str:
    text = str(value).strip().lower()
    if text in {"1", "1.0", "true", "spatial"}:
        return SPATIAL
    if text in {"0", "0.0", "false", "non-spatial", "nonspatial"}:
        return NON_SPATIAL
    return text or NON_SPATIAL


def _ordered_values(series: pd.Series) -> list[str]:
    working = series.fillna("").astype(str).str.strip()
    counts = working.value_counts()
    return counts.index.tolist()


def _count_pivot(
    frame: pd.DataFrame,
    *,
    index: str,
    columns: str,
    column_order: Optional[list[str]] = None,
    index_order: Optional[list[str]] = None,
    sort_index: bool = False,
) -> pd.DataFrame:
    if frame.empty:
        result = pd.DataFrame(columns=[index])
        return result

    pivot = pd.pivot_table(
        frame,
        index=index,
        columns=columns,
        values=frame.columns[0],
        aggfunc="count",
        fill_value=0,
    )
    pivot.columns.name = None
    if column_order is not None:
        ordered = [column for column in column_order if column in pivot.columns]
        pivot = pivot.reindex(columns=ordered, fill_value=0)
    if index_order is not None:
        ordered_index = [value for value in index_order if value in pivot.index]
        pivot = pivot.reindex(ordered_index, fill_value=0)
    if sort_index:
        pivot = pivot.sort_index()
    result = pivot.reset_index()
    return result.rename(columns={index: DISPLAY_COLUMN_NAMES.get(index, index)})


def _with_totals(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    working = frame.copy()
    id_col = working.columns[0]
    value_cols = list(working.columns[1:])
    working[TOTAL_LABEL] = working[value_cols].sum(axis=1)
    total_row = {id_col: TOTAL_LABEL}
    for column in value_cols + [TOTAL_LABEL]:
        total_row[column] = working[column].sum()
    return pd.concat([working, pd.DataFrame([total_row])], ignore_index=True)


def _share_table(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    working = frame.copy()
    id_col = working.columns[0]
    value_cols = list(working.columns[1:])
    row_totals = working[value_cols].sum(axis=1).replace(0, np.nan)
    share_values = working[value_cols].div(row_totals, axis=0).fillna(0.0)
    result = pd.concat([working[[id_col]], share_values], axis=1)
    result[TOTAL_LABEL] = result[value_cols].sum(axis=1)

    overall = working[value_cols].sum(axis=0)
    grand_total = overall.sum()
    total_row = {id_col: TOTAL_LABEL}
    for column in value_cols:
        total_row[column] = overall[column] / grand_total if grand_total else 0.0
    total_row[TOTAL_LABEL] = 1.0 if grand_total else 0.0
    return pd.concat([result, pd.DataFrame([total_row])], ignore_index=True)
