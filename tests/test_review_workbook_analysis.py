from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from src.review_workbook_analysis import (
    ANALYSIS_SHEET_NAME,
    ANALYSIS_SHEETS,
    ANALYSIS_SPATIAL_LEVEL,
    CHART_SHEET_NAME,
    COL_SPACE_DETAIL,
    COL_SPATIAL_LEVEL,
    COL_TOPIC_NAME_ZH,
    COL_YEAR,
    KPI_SHEET_NAME,
    LEVEL_SHARE_SHEET,
    LEVEL_TOPIC_COUNT_SHEET,
    NOT_MENTIONED,
    SPACE_COUNT_SHEET,
    SOURCE_SHEET_NAME,
    TOTAL_LABEL,
    TOPIC_SPACE_TOP_SHEET,
    YEAR_SPATIAL_FLAG_SHEET,
    YEAR_TOPIC_COUNT_SHEET,
    generate_review_analysis_workbook,
)


def _build_sample_review_frame() -> pd.DataFrame:
    spatial_flag = "\u9884\u6d4b_\u7a7a\u95f4\u7814\u7a76/\u975e\u7a7a\u95f4\u7814\u7a76"
    spatial_level = "\u9884\u6d4b_\u7a7a\u95f4\u7b49\u7ea7"
    space_detail = "\u9884\u6d4b_\u5177\u4f53\u7a7a\u95f4\u63cf\u8ff0"
    urban_flag = "\u9884\u6d4b_\u662f\u5426\u5c5e\u4e8e\u57ce\u5e02\u66f4\u65b0\u7814\u7a76"
    urban_confidence = "\u57ce\u5e02\u66f4\u65b0\u5224\u5b9a\u7f6e\u4fe1\u5ea6(confidence)"
    reasoning = "\u7a7a\u95f4\u63d0\u53d6\u4f9d\u636e(Reasoning)"
    spatial_confidence = "\u7a7a\u95f4\u63d0\u53d6\u7f6e\u4fe1\u5ea6(Confidence)"

    return pd.DataFrame(
        [
            {
                "Article Title": "Paper A",
                COL_YEAR: 2020,
                "Keywords Plus": "renewal",
                "Abstract": "A",
                "WoS Categories": "Urban Studies",
                "Research Areas": "Geography",
                urban_flag: "1",
                spatial_flag: "1",
                spatial_level: "7. Single-city / Municipal Scale",
                space_detail: "Shanghai",
                "topic_final": "U1",
                "topic_final_name_en": "old neighborhood renewal",
                "topic_final_name_zh": "\u8001\u65e7\u4f4f\u533a\u66f4\u65b0",
                urban_confidence: 0.95,
                reasoning: "district level evidence",
                spatial_confidence: "High",
                "review_flag": 0,
                "review_reason": "",
            },
            {
                "Article Title": "Paper B",
                COL_YEAR: 2020,
                "Keywords Plus": "renewal",
                "Abstract": "B",
                "WoS Categories": "Urban Studies",
                "Research Areas": "Geography",
                urban_flag: "1",
                spatial_flag: "1",
                spatial_level: "9. Micro / Neighborhood / Block",
                space_detail: "Shanghai",
                "topic_final": "U1",
                "topic_final_name_en": "old neighborhood renewal",
                "topic_final_name_zh": "\u8001\u65e7\u4f4f\u533a\u66f4\u65b0",
                urban_confidence: 0.91,
                reasoning: "block level evidence",
                spatial_confidence: "High",
                "review_flag": 0,
                "review_reason": "",
            },
            {
                "Article Title": "Paper C",
                COL_YEAR: 2021,
                "Keywords Plus": "gentrification",
                "Abstract": "C",
                "WoS Categories": "Urban Studies",
                "Research Areas": "Sociology",
                urban_flag: "1",
                spatial_flag: "1",
                spatial_level: "8. District / County Scale",
                space_detail: "Shenzhen",
                "topic_final": "U12",
                "topic_final_name_en": "gentrification exclusion and neighborhood change",
                "topic_final_name_zh": "\u7ec5\u58eb\u5316\u3001\u6392\u65a5\u4e0e\u793e\u533a\u53d8\u5316",
                urban_confidence: 0.89,
                reasoning: "district scale evidence",
                spatial_confidence: "High",
                "review_flag": 0,
                "review_reason": "",
            },
            {
                "Article Title": "Paper D",
                COL_YEAR: 2021,
                "Keywords Plus": "governance",
                "Abstract": "D",
                "WoS Categories": "Urban Studies",
                "Research Areas": "Policy",
                urban_flag: "0",
                spatial_flag: "0",
                spatial_level: NOT_MENTIONED,
                space_detail: NOT_MENTIONED,
                "topic_final": "N3",
                "topic_final_name_en": "general urban governance",
                "topic_final_name_zh": "\u4e00\u822c\u57ce\u5e02\u6cbb\u7406",
                urban_confidence: 0.67,
                reasoning: "",
                spatial_confidence: "",
                "review_flag": 0,
                "review_reason": "",
            },
            {
                "Article Title": "Paper E",
                COL_YEAR: 2022,
                "Keywords Plus": "policy",
                "Abstract": "E",
                "WoS Categories": "Urban Studies",
                "Research Areas": "Economics",
                urban_flag: "1",
                spatial_flag: "1",
                spatial_level: "3. National / Single-country Scale",
                space_detail: "Beijing",
                "topic_final": "U10",
                "topic_final_name_en": "renewal finance and policy tools",
                "topic_final_name_zh": "\u66f4\u65b0\u878d\u8d44\u4e0e\u653f\u7b56\u5de5\u5177",
                urban_confidence: 0.93,
                reasoning: "national evidence",
                spatial_confidence: "Medium",
                "review_flag": 0,
                "review_reason": "",
            },
            {
                "Article Title": "Paper F",
                COL_YEAR: "",
                "Keywords Plus": "policy",
                "Abstract": "F",
                "WoS Categories": "Urban Studies",
                "Research Areas": "Economics",
                urban_flag: "1",
                spatial_flag: "1",
                spatial_level: "6. Multi-city / Megaregion Scale",
                space_detail: "Guangzhou",
                "topic_final": "U10",
                "topic_final_name_en": "renewal finance and policy tools",
                "topic_final_name_zh": "\u66f4\u65b0\u878d\u8d44\u4e0e\u653f\u7b56\u5de5\u5177",
                urban_confidence: 0.88,
                reasoning: "multi-city evidence",
                spatial_confidence: "Medium",
                "review_flag": 0,
                "review_reason": "",
            },
            {
                "Article Title": "Paper G",
                COL_YEAR: 2022,
                "Keywords Plus": "unknown",
                "Abstract": "G",
                "WoS Categories": "Urban Studies",
                "Research Areas": "Geography",
                urban_flag: "1",
                spatial_flag: "1",
                spatial_level: "9. Micro / Neighborhood / Block Scale",
                space_detail: NOT_MENTIONED,
                "topic_final": "Unknown",
                "topic_final_name_en": "Unknown",
                "topic_final_name_zh": "\u672a\u77e5\u4e3b\u9898",
                urban_confidence: 0.61,
                reasoning: "unclear scale",
                spatial_confidence: "Low",
                "review_flag": 1,
                "review_reason": "manual check",
            },
        ]
    )


def test_generate_review_analysis_workbook_appends_analysis_sheets(tmp_path):
    workbook_path = tmp_path / "review.xlsx"
    source_df = _build_sample_review_frame()
    source_df.to_excel(workbook_path, sheet_name=SOURCE_SHEET_NAME, index=False, engine="openpyxl")
    original_df = pd.read_excel(workbook_path, sheet_name=SOURCE_SHEET_NAME, engine="openpyxl")

    generate_review_analysis_workbook(
        workbook_path,
        append=True,
        replace_analysis_sheets=True,
    )
    generate_review_analysis_workbook(
        workbook_path,
        append=True,
        replace_analysis_sheets=True,
    )

    workbook = load_workbook(workbook_path)
    assert SOURCE_SHEET_NAME in workbook.sheetnames
    for sheet_name in ANALYSIS_SHEETS:
        assert sheet_name in workbook.sheetnames

    roundtrip_df = pd.read_excel(workbook_path, sheet_name=SOURCE_SHEET_NAME, engine="openpyxl")
    pd.testing.assert_frame_equal(roundtrip_df, original_df)

    year_topic_count = pd.read_excel(workbook_path, sheet_name=YEAR_TOPIC_COUNT_SHEET, engine="openpyxl")
    value_cols = [column for column in year_topic_count.columns[1:] if column != TOTAL_LABEL]
    body = year_topic_count[year_topic_count[COL_YEAR] != TOTAL_LABEL]
    assert int(body[value_cols].to_numpy().sum()) == 6

    year_spatial_flag = pd.read_excel(workbook_path, sheet_name=YEAR_SPATIAL_FLAG_SHEET, engine="openpyxl")
    by_year = year_spatial_flag[year_spatial_flag[COL_YEAR] != TOTAL_LABEL]
    assert (by_year[["Spatial", "Non-spatial"]].sum(axis=1) == by_year["Total_Count"]).all()

    level_share = pd.read_excel(workbook_path, sheet_name=LEVEL_SHARE_SHEET, engine="openpyxl")
    assert int(level_share["Count"].sum()) == 6

    space_count = pd.read_excel(workbook_path, sheet_name=SPACE_COUNT_SHEET, engine="openpyxl")
    assert int(space_count["Count"].sum()) == 5

    analysis_df = pd.read_excel(workbook_path, sheet_name=ANALYSIS_SHEET_NAME, engine="openpyxl")
    assert "9. Micro / Neighborhood / Block" not in analysis_df[ANALYSIS_SPATIAL_LEVEL].tolist()
    assert "9. Micro / Neighborhood / Block Scale" in analysis_df[ANALYSIS_SPATIAL_LEVEL].tolist()

    charts_ws = workbook[CHART_SHEET_NAME]
    assert len(charts_ws._images) >= 6
    assert workbook[KPI_SHEET_NAME]["A1"].value == "Summary Metrics"


def test_level_topic_and_space_tables_are_written(tmp_path):
    workbook_path = tmp_path / "review.xlsx"
    _build_sample_review_frame().to_excel(workbook_path, sheet_name=SOURCE_SHEET_NAME, index=False, engine="openpyxl")

    generate_review_analysis_workbook(
        workbook_path,
        append=True,
        replace_analysis_sheets=True,
    )

    level_topic = pd.read_excel(workbook_path, sheet_name=LEVEL_TOPIC_COUNT_SHEET, engine="openpyxl")
    assert COL_SPATIAL_LEVEL in level_topic.columns
    assert COL_TOPIC_NAME_ZH in pd.read_excel(
        workbook_path,
        sheet_name=TOPIC_SPACE_TOP_SHEET,
        engine="openpyxl",
    ).columns

    assert COL_SPACE_DETAIL in pd.read_excel(
        workbook_path,
        sheet_name=SPACE_COUNT_SHEET,
        engine="openpyxl",
    ).columns

    assert level_topic.iloc[-1, 0] == TOTAL_LABEL
